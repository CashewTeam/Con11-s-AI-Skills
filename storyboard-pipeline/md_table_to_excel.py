import re
import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


COLOR_MAP = {
    "实拍": "FFCCCC",
    "Aroll": "CCFFCC",
    "AROLL": "CCFFCC",
    "aroll": "CCFFCC",
    "动效": "CCCCFF",
    "资料": "CCE5FF",
}

TYPE_HEADER_KEYWORDS = ["镜头类型", "镜头类别", "type", "shot type"]

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
header_font = Font(bold=True, size=11)
cell_font = Font(size=11)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_alignment = Alignment(vertical="center", wrap_text=True)


def find_table_section(lines: list[str]) -> tuple[int, int] | None:
    separator_idx = None
    for i, line in enumerate(lines):
        stripped = line.strip()
        if stripped.startswith("|") and "---" in stripped and stripped.count("|") > 1:
            separator_idx = i
            break

    if separator_idx is None or separator_idx == 0:
        return None

    header_idx = separator_idx - 1
    header_line = lines[header_idx].strip()
    if not header_line.startswith("|"):
        return None

    col_count = header_line.count("|") - 1

    end_idx = separator_idx + 1
    while end_idx < len(lines):
        row = lines[end_idx].strip()
        if row.startswith("|") and row.count("|") - 1 == col_count:
            end_idx += 1
        else:
            break

    return header_idx, end_idx


def parse_cells(line: str) -> list[str]:
    return [cell.strip() for cell in line.strip().split("|")[1:-1]]


def detect_type_column_index(headers: list[str]) -> int | None:
    for i, h in enumerate(headers):
        for kw in TYPE_HEADER_KEYWORDS:
            if kw.lower() in h.lower():
                return i
    return None


def get_color_for_type(value: str) -> str | None:
    if not value:
        return None
    for keyword, color in COLOR_MAP.items():
        if value.startswith(keyword):
            return color
    primary = value.split("·")[0].strip() if "·" in value else value.strip()
    return COLOR_MAP.get(primary)


def parse_markdown_table(text: str) -> tuple[list[str], list[list[str]]]:
    lines = text.splitlines()
    result = find_table_section(lines)
    if result is None:
        raise ValueError("未在输入中找到有效的 Markdown 表格")
    header_idx, end_idx = result
    headers = parse_cells(lines[header_idx])
    rows = []
    for i in range(header_idx + 2, end_idx):
        row = lines[i].strip()
        if row.startswith("|"):
            rows.append(parse_cells(row))
    return headers, rows


def write_excel(
    headers: list[str],
    rows: list[list[str]],
    output_path: Path,
    type_col_index: int,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "分镜脚本"

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border

            if col_idx == type_col_index + 1:
                color = get_color_for_type(value)
                if color:
                    cell.fill = PatternFill(
                        start_color=color, end_color=color, fill_type="solid"
                    )

    col_widths = {}
    for col_idx in range(1, len(headers) + 1):
        max_len = 0
        for row_idx in range(1, len(rows) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                chinese_count = sum(1 for c in str(val) if "\u4e00" <= c <= "\u9fff")
                ascii_count = len(str(val)) - chinese_count
                effective_len = chinese_count * 2.2 + ascii_count * 1.1
                max_len = max(max_len, effective_len)
        col_widths[col_idx] = min(max_len + 4, 45)

    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = max(width, 10)

    ws.freeze_panes = "A2"
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description="将 Markdown 表格转换为带彩色标注的 Excel 文件"
    )
    parser.add_argument(
        "input",
        nargs="?",
        help="输入的 Markdown 文件路径，不提供则从标准输入读取",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="storyboard.xlsx",
        help="输出的 Excel 文件路径 (默认: storyboard.xlsx)",
    )
    args = parser.parse_args()

    if args.input:
        text = Path(args.input).read_text(encoding="utf-8")
    else:
        import sys

        print("请输入 Markdown 表格内容，按 Ctrl+Z 后回车结束:")
        text = sys.stdin.read()

    headers, rows = parse_markdown_table(text)

    type_col_index = detect_type_column_index(headers)
    if type_col_index is None:
        print('警告: 未找到"镜头类型"列，将不应用颜色标注')

    output_path = Path(args.output)
    write_excel(headers, rows, output_path, type_col_index)
    print(f"已生成 Excel 文件: {output_path.resolve()}")


if __name__ == "__main__":
    main()
